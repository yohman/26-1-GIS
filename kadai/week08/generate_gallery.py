#!/usr/bin/env python3
"""
Generate the Week 08 midterm gallery from the final group CSV and the
submission workbook.

Run this from anywhere:
    python kadai/week08/generate_gallery.py
"""

from __future__ import annotations

import csv
import html
import re
import unicodedata
from html.parser import HTMLParser
from urllib.error import URLError, HTTPError
from urllib.request import Request, urlopen
from pathlib import Path

from openpyxl import load_workbook


BASE_DIR = Path(__file__).resolve().parent
GROUPS_CSV = BASE_DIR / "Final GIS Groups.csv"
WORKBOOK = BASE_DIR / "midterms.xlsx"
OUTPUT_HTML = BASE_DIR / "index.html"

COUNTRY_COLORS = {
    "Japan": "#D00027",
    "Brazil": "#009B3A",
    "France": "#0055A4",
    "Argentina": "#5DADEC",
    "Spain": "#C60B1E",
    "Germany": "#1F1F1F",
    "Italy": "#008C45",
    "Portugal": "#006600",
    "Netherlands": "#FF6F00",
    "England": "#B31942",
    "Mexico": "#006847",
    "Korea": "#0047A0",
    "Morocco": "#C1272D",
    "Croatia": "#171796",
    "USA": "#3C3B6E",
    "Australia": "#0057B8",
    "Senegal": "#00853F",
    "Uruguay": "#0038A8",
    "Ghana": "#FCD116",
    "Canada": "#D52B1E",
    "Colombia": "#FCD116",
}


def normalize_text(value: object) -> str:
    return unicodedata.normalize("NFKC", str(value or "")).strip()


def split_name(value: object) -> tuple[str, str]:
    text = normalize_text(value)
    if not text:
        return "", ""

    if "（" in text and text.endswith("）"):
        name, kana = text.split("（", 1)
        return name.strip(), kana.rstrip("）").strip()
    if "(" in text and text.endswith(")"):
        name, kana = text.split("(", 1)
        return name.strip(), kana.rstrip(")").strip()
    return text, ""


def load_groups(path: Path) -> list[dict]:
    with path.open("r", encoding="utf-8-sig", newline="") as f:
        reader = csv.DictReader(f)
        rows = list(reader)

    groups: dict[str, dict] = {}
    for row in rows:
        group_number = str(row["Group"]).strip()
        group = groups.setdefault(
            group_number,
            {
                "group_number": int(group_number),
                "country": row["Country"].strip(),
                "code": row["Country Code"].strip(),
                "color": COUNTRY_COLORS.get(row["Country"].strip(), "#475569"),
                "members": [],
            },
        )
        group["members"].append(
            {
                "id": normalize_text(row["学籍番号"]),
                "csv_name": normalize_text(row["Name"]),
                "csv_kana": normalize_text(row["Kana Name"]),
                "english": normalize_text(row["English Name"]),
                "department": normalize_text(row["Department"]),
                "missed": normalize_text(row["Classes Missed"]),
                "affiliation": normalize_text(row["Affiliation"]),
                "failing_risk": normalize_text(row["Failing Risk"]).lower() == "yes",
            }
        )

    return [groups[key] for key in sorted(groups, key=lambda value: int(value))]


def clean_url(url: str) -> str:
    cleaned = url.strip().strip("\"'<>[]{}")

    # Some workbook cells append notes directly after the URL with no space.
    for marker in ("(padlet", "(gitpages", "(github", "(github pages", "（padlet", "（gitpages", "（github"):
        idx = cleaned.lower().find(marker.lower())
        if idx != -1:
            cleaned = cleaned[:idx]

    cleaned = cleaned.rstrip(".,;:!?)]}>'\"")
    return cleaned.strip()


URL_RE = re.compile(r"https?://[^\s<>\"']+")


class TitleParser(HTMLParser):
    def __init__(self) -> None:
        super().__init__()
        self.in_title = False
        self.title_parts: list[str] = []

    def handle_starttag(self, tag, attrs):
        if tag.lower() == "title":
            self.in_title = True

    def handle_endtag(self, tag):
        if tag.lower() == "title":
            self.in_title = False

    def handle_data(self, data):
        if self.in_title:
            self.title_parts.append(data)

    def title(self) -> str:
        return normalize_text("".join(self.title_parts))


def extract_links(*chunks: object) -> dict[str, list[str]]:
    links: dict[str, list[str]] = {
        "padlet": [],
        "github_pages": [],
        "github_repo": [],
        "other": [],
    }
    text = "\n".join(normalize_text(chunk) for chunk in chunks if normalize_text(chunk))

    for raw_url in URL_RE.findall(text):
        url = clean_url(raw_url)
        lower = url.lower()
        if "padlet.com" in lower:
            links["padlet"].append(url)
        elif "github.io" in lower:
            links["github_pages"].append(url)
        elif "github.com" in lower:
            links["github_repo"].append(url)
        else:
            links["other"].append(url)

    return links


def dedupe_urls(urls: list[str]) -> list[str]:
    seen: set[str] = set()
    deduped: list[str] = []
    for url in urls:
        if url not in seen:
            seen.add(url)
            deduped.append(url)
    return deduped


def fetch_page_title(url: str) -> str:
    try:
        request = Request(url, headers={"User-Agent": "Mozilla/5.0"})
        with urlopen(request, timeout=8) as response:
            content_type = response.headers.get_content_charset() or "utf-8"
            html_text = response.read().decode(content_type, errors="replace")
    except (HTTPError, URLError, TimeoutError, ValueError, UnicodeError, OSError):
        return ""

    parser = TitleParser()
    parser.feed(html_text)
    title = parser.title()
    if title:
        return title

    match = re.search(r'property=["\']og:title["\'][^>]*content=["\']([^"\']+)', html_text, re.I)
    if match:
        return normalize_text(match.group(1))
    return ""


def slug_title_from_url(url: str) -> str:
    cleaned = re.sub(r"^https?://", "", url, flags=re.I)
    cleaned = cleaned.split("?", 1)[0].split("#", 1)[0]
    cleaned = cleaned.strip("/")
    parts = cleaned.split("/")
    slug = parts[-1] if parts else cleaned
    slug = slug.replace("-", " ").replace("_", " ").replace(".", " ")
    slug = re.sub(r"\s+", " ", slug).strip()
    if not slug:
        return "Untitled Project"
    return slug.title()


def infer_project_title(links: dict[str, list[str]]) -> str:
    candidates = dedupe_urls(links["github_pages"] + links["other"] + links["github_repo"])
    for url in candidates:
        title = fetch_page_title(url)
        if title:
            return title
    if candidates:
        return slug_title_from_url(candidates[0])
    if links["padlet"]:
        return "Padlet Project"
    return "Untitled Project"


def load_submissions(path: Path) -> dict[str, dict]:
    workbook = load_workbook(path, data_only=False)
    sheet = workbook.active

    submissions: dict[str, dict] = {}
    for row in sheet.iter_rows(min_row=2, values_only=True):
        student_id = normalize_text(row[0])
        if not student_id:
            continue

        csv_name, csv_kana = split_name(row[1])
        links = extract_links(row[8] if len(row) > 8 else "", row[9] if len(row) > 9 else "")
        project_title = infer_project_title(links)

        submissions[student_id] = {
            "workbook_name": csv_name,
            "workbook_kana": csv_kana,
            "links": links,
            "project_title": project_title,
            "submitted_at": normalize_text(row[6] if len(row) > 6 else ""),
        }

    return submissions


def student_card_html(member: dict, submission: dict | None) -> str:
    name = member["csv_name"] or (submission["workbook_name"] if submission else "")
    kana = member["csv_kana"] or (submission["workbook_kana"] if submission else "")
    links = submission["links"] if submission else {}
    project_title = submission["project_title"] if submission else "Untitled Project"
    untitled = project_title == "Untitled Project"

    link_bits = []
    for url in links.get("padlet", []):
        link_bits.append(
            f'<a class="project-link padlet" href="{html.escape(url, quote=True)}" target="_blank" rel="noopener">Padlet</a>'
        )
    for url in links.get("github_pages", []):
        link_bits.append(
            f'<a class="project-link github" href="{html.escape(url, quote=True)}" target="_blank" rel="noopener">GitHub Pages</a>'
        )
    for url in links.get("github_repo", []):
        link_bits.append(
            f'<a class="project-link repo" href="{html.escape(url, quote=True)}" target="_blank" rel="noopener">GitHub Repo</a>'
        )
    for url in links.get("other", []):
        host = re.sub(r"^https?://", "", url, flags=re.I).split("/", 1)[0]
        link_bits.append(
            f'<a class="project-link other" href="{html.escape(url, quote=True)}" target="_blank" rel="noopener">{html.escape(host)}</a>'
        )

    if link_bits:
        link_html = '<div class="project-links">' + "".join(link_bits) + "</div>"
    else:
        link_html = '<div class="project-links empty">No project links yet</div>'

    return f"""
        <div class="member{ ' risk' if member['failing_risk'] else '' }{ ' untitled' if untitled else '' }">
            <div class="project-title">{html.escape(project_title)}</div>
            <div class="member-top">
                <div class="name">{html.escape(name)}</div>
                <div class="badge { 'eng' if member['affiliation'] == 'Engineer' else 'non' }{ ' risk' if member['failing_risk' ] else '' }">{'ENG' if member['affiliation'] == 'Engineer' else 'NON'}</div>
            </div>
            <div class="kana">{html.escape(kana)}</div>
            {link_html}
        </div>
    """


def render_html(groups: list[dict], submissions: dict[str, dict]) -> str:
    total_students = sum(len(group["members"]) for group in groups)
    linked_students = sum(1 for group in groups for member in group["members"] if submissions.get(member["id"], {}).get("links"))
    padlet_count = sum(len(submissions.get(member["id"], {}).get("links", {}).get("padlet", [])) for group in groups for member in group["members"])
    github_count = sum(len(submissions.get(member["id"], {}).get("links", {}).get("github_pages", [])) for group in groups for member in group["members"])
    extra_count = sum(
        len(submissions.get(member["id"], {}).get("links", {}).get("github_repo", []))
        + len(submissions.get(member["id"], {}).get("links", {}).get("other", []))
        for group in groups
        for member in group["members"]
    )

    group_cards = []
    for group in groups:
        ordered_members = sorted(
            group["members"],
            key=lambda member: (
                0 if member["id"] in submissions else 1,
                1 if member["failing_risk"] else 0,
                member["csv_name"] or member["english"] or member["id"],
            ),
        )
        members_html = "".join(
            student_card_html(member, submissions.get(member["id"]))
            for member in ordered_members
        )
        group_cards.append(
            f"""
            <article class="group-card" style="--team-color: {group['color']}">
                <div class="group-band">
                    <div class="group-title">GROUP {group['group_number']:02d}</div>
                    <div class="group-code">{html.escape(group['code'])}</div>
                </div>
                <div class="group-members">
                    {members_html}
                </div>
            </article>
            """
        )

    return f"""<!DOCTYPE html>
<html lang="en">

<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>2026 Reitaku GIS Midterm Presentations</title>
    <style>
        :root {{
            --bg: #ffffff;
            --panel: #ffffff;
            --ink: #111827;
            --muted: #4b5563;
            --line: #d1d5db;
            --soft: #fafafa;
            --risk-bg: #f3f4f6;
            --risk: #6b7280;
        }}

        * {{
            box-sizing: border-box;
        }}

        body {{
            margin: 0;
            font-family: Arial, Helvetica, sans-serif;
            background: var(--bg);
            color: var(--ink);
        }}

        .app {{
            min-height: 100vh;
        }}

        .topbar {{
            position: sticky;
            top: 0;
            z-index: 20;
            background: rgba(255, 255, 255, 0.98);
            backdrop-filter: blur(6px);
            padding: 18px 20px 14px;
        }}

        .topbar-inner {{
            display: flex;
            align-items: stretch;
            justify-content: space-between;
            gap: 24px;
        }}

        .title-block {{
            flex: 1 1 auto;
            align-self: center;
        }}

        .title-block h1 {{
            margin: 0;
            font-size: 17px;
            line-height: 1.1;
            text-transform: uppercase;
            letter-spacing: 0.1em;
        }}

        .title-block p {{
            margin: 6px 0 0;
            color: var(--muted);
            font-size: 12px;
            line-height: 1.35;
        }}

        .summary {{
            display: flex;
            align-items: center;
            gap: 10px;
            flex-wrap: wrap;
            padding: 16px 0 4px;
        }}

        .summary-chip {{
            border: none;
            background: #f8fafc;
            border-radius: 0;
            padding: 7px 10px;
            font-size: 11px;
            color: var(--muted);
        }}

        .view {{
            padding: 18px 20px 28px;
        }}

        .section-title {{
            margin: 0 0 14px;
            font-size: 11px;
            font-weight: 800;
            text-transform: uppercase;
            letter-spacing: 0.18em;
            color: var(--muted);
        }}

        .groups-grid {{
            display: grid;
            grid-template-columns: repeat(auto-fit, minmax(300px, 1fr));
            gap: 16px;
        }}

        .group-card {{
            border: none;
            border-radius: 0;
            background: white;
        }}

        .group-band {{
            background: var(--team-color);
            color: white;
            padding: 10px 12px;
            display: flex;
            align-items: center;
            justify-content: space-between;
            gap: 10px;
        }}

        .group-title {{
            font-size: 10px;
            font-weight: 800;
            letter-spacing: 0.18em;
        }}

        .group-code {{
            font-size: 16px;
            font-weight: 900;
            letter-spacing: 0.06em;
        }}

        .group-members {{
            padding: 0;
            display: grid;
            gap: 10px;
            margin-top: 10px;
        }}

        .member {{
            border-left: 8px solid var(--team-color);
            border: 1px solid #e5e7eb;
            border-radius: 0;
            padding: 10px 10px 10px 12px;
            background: white;
        }}

        .member.risk {{
            background: #f4f5f7;
            border-left-color: var(--risk);
            border-color: #d1d5db;
            color: #6b7280;
        }}

        .member.untitled {{
            background: #f8fafc;
            border-left-color: #e2e8f0;
            border-color: #e2e8f0;
            color: #94a3b8;
        }}

        .member-top {{
            display: flex;
            align-items: center;
            justify-content: space-between;
            gap: 8px;
        }}

        .name {{
            font-size: 14px;
            line-height: 1.15;
            font-weight: 800;
        }}

        .kana {{
            margin-top: 3px;
            font-size: 10px;
            color: var(--muted);
        }}

        .project-title {{
            margin-bottom: 8px;
            font-size: 20px;
            line-height: 1.25;
            font-weight: 900;
            letter-spacing: 0.05em;
            text-transform: uppercase;
            color: var(--team-color);
            min-height: 36px;
        }}

        .member.untitled .project-title {{
            color: #94a3b8;
        }}

        .badge {{
            flex: 0 0 auto;
            border: none;
            border-radius: 0;
            color: white;
            font-size: 9px;
            font-weight: 900;
            padding: 4px 6px;
            letter-spacing: 0.05em;
            position: relative;
        }}

        .badge.eng {{
            background: #111827;
            border-color: #111827;
            color: white;
            opacity: 0.2;
        }}

        .badge.non {{
            background: white;
            border: 1px solid #111827;
            color: #111827;
            opacity: 0.2;
        }}

        .member.risk .badge {{
            background: #111827;
            border-color: #111827;
            color: white;
        }}

        .badge.risk::after {{
            content: "";
            position: absolute;
            left: 3px;
            right: 3px;
            top: 50%;
            border-top: 1px solid rgba(255,255,255,0.95);
            transform: rotate(-18deg);
            transform-origin: center;
            pointer-events: none;
        }}

        .project-links {{
            display: flex;
            flex-wrap: wrap;
            gap: 8px;
            margin-top: 8px;
        }}

        .project-link {{
            display: inline-flex;
            align-items: center;
            justify-content: center;
            border-radius: 0;
            padding: 7px 10px;
            font-size: 10px;
            font-weight: 800;
            letter-spacing: 0.08em;
            text-transform: uppercase;
            text-decoration: none;
            border: 1px solid #111827;
            color: #111827;
            background: white;
        }}

        .project-link.padlet {{
            background: #111827;
            color: white;
            border-color: #111827;
        }}

        .project-link.github {{
            background: white;
            color: #111827;
        }}

        .project-link.repo,
        .project-link.other {{
            background: #f8fafc;
            color: #111827;
            border-color: #d1d5db;
        }}

        .project-links.empty {{
            margin-top: 8px;
            font-size: 10px;
            color: var(--muted);
        }}

        footer {{
            text-align: center;
            padding: 0 20px 18px;
            color: var(--muted);
            font-size: 11px;
        }}

        @media (max-width: 760px) {{
            .groups-grid {{
                grid-template-columns: 1fr;
            }}
        }}
    </style>
</head>

<body>
    <div class="app">
        <div class="topbar">
            <div class="topbar-inner">
                <div class="title-block">
                    <h1>2026 Reitaku GIS Midterm Presentations</h1>
                </div>
            </div>
            <div class="summary">
                <div class="summary-chip">Groups: {len(groups)}</div>
                <div class="summary-chip">Students: {total_students}</div>
                <div class="summary-chip">With links: {linked_students}</div>
                <div class="summary-chip">Padlet: {padlet_count}</div>
                <div class="summary-chip">GitHub Pages: {github_count}</div>
                <div class="summary-chip">Extras: {extra_count}</div>
            </div>
        </div>

        <section class="view">
            <h2 class="section-title">Final Groups</h2>
            <div class="groups-grid">
                {''.join(group_cards)}
            </div>
        </section>
    </div>

    <footer>
        Generated from Final GIS Groups.csv and midterms.xlsx. Re-run generate_gallery.py after updates.
    </footer>
</body>

</html>
"""


def main() -> None:
    if not GROUPS_CSV.exists():
        raise FileNotFoundError(f"Missing groups CSV: {GROUPS_CSV}")
    if not WORKBOOK.exists():
        raise FileNotFoundError(f"Missing workbook: {WORKBOOK}")

    groups = load_groups(GROUPS_CSV)
    submissions = load_submissions(WORKBOOK)
    OUTPUT_HTML.write_text(render_html(groups, submissions), encoding="utf-8")
    print(f"Wrote {OUTPUT_HTML}")


if __name__ == "__main__":
    main()
